import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np

some_even_more_new_variable = 1


class DataSet:
    """
    Класс для представления датасета
    Attributes:
        file_name(Any): Названия файла датасета
        vacancy_name(Any): Название вакансии
    """
    def __init__(self, file_name, vacancy_name):
        """
        Инициализирует объект DataSet, считает общую зарплату, считает среднее значение зарплаты,
        считывает значение из файла, считает статистику, выводит значение получившиеся значения на консоль

        Args:
            file_name(Any): Названия файла датасета
            vacancy_name(Any): Название вакансии
        """
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def add_amount_salary(salary_dict, key, amount_of_salary):
        """
        Вычисляет общуюю сумму зарплат
        :param
            salary_dict: словарь зарплат
        :param
            key: ключ для зарплаты по вакансиям
        :param
            amount_of_salary: нужное количество добавки зарплаты
        :return:
            void
        """
        if key in salary_dict:
            salary_dict[key] += amount_of_salary
        else:
            salary_dict[key] = amount_of_salary

    @staticmethod
    def get_average_salary(salary_dict):
        """
        Считает словарь средних зарплат
        :param
            salary_dict: словарь зарплат
        :return:
            dict: словарь средних зарплат
        """
        new_dictionary = {}
        for key, values in salary_dict.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def work_with_the_file(self):
        """
        Считывает значения с файла
        :return:
            void
        """
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader_csv = csv.reader(file)
            header_of_the_file = next(reader_csv)
            header_length = len(header_of_the_file)
            for row_in_file in reader_csv:
                if '' not in row_in_file and len(row_in_file) == header_length:
                    yield dict(zip(header_of_the_file, row_in_file))

    def calculate_statistics(self):
        """
        Вычисляет статистику:
            Динамика уровня зарплат по годам,
            Динамика количества вакансий по годам,
            Динамика уровня зарплат по годам для выбранной профессии,
            Динамика количества вакансий по годам для выбранной професси,
            Уровень зарплат по городам,
            Доля вакансий по городам

        :return:
            void
        """
        salary = {}
        salary_of_vacancy_name = {}
        salary_city = {}
        count_of_vacancies = 0

        for vacancies_dict in self.work_with_the_file():
            vacancy = Salary(vacancies_dict)
            self.add_amount_salary(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.add_amount_salary(salary_of_vacancy_name, vacancy.year, [vacancy.salary_average])
            self.add_amount_salary(salary_city, vacancy.area_name, [vacancy.salary_average])
            count_of_vacancies += 1

        vac_num_dict = dict([(key, len(value)) for key, value in salary.items()])
        vac_by_name = dict([(key, len(value)) for key, value in salary_of_vacancy_name.items()])

        if not salary_of_vacancy_name:
            salary_of_vacancy_name = dict([(key, [0]) for key, value in salary.items()])
            vac_by_name = dict([(key, 0) for key, value in vac_num_dict.items()])
        average_salary = self.get_average_salary(salary)
        average_salary_vac = self.get_average_salary(salary_of_vacancy_name)
        average_salary_city = self.get_average_salary(salary_city)
        quantity_dynamics = {}

        for year, salaries in salary_city.items():
            quantity_dynamics[year] = round(len(salaries) / count_of_vacancies, 4)
        quantity_dynamics = list(filter(lambda a: a[-1] >= 0.01,
                                        [(key, value) for key, value in quantity_dynamics.items()]))
        quantity_dynamics.sort(key=lambda a: a[-1], reverse=True)
        top_ten_quantity = quantity_dynamics.copy()
        quantity_dynamics = dict(quantity_dynamics)
        average_salary_city = list(filter(lambda a: a[0] in list(quantity_dynamics.keys()),
                                          [(key, value) for key, value in average_salary_city.items()]))
        average_salary_city.sort(key=lambda a: a[-1], reverse=True)
        average_salary_city = dict(average_salary_city[:10])
        top_ten_quantity = dict(top_ten_quantity[:10])

        return average_salary, vac_num_dict, average_salary_vac, vac_by_name, average_salary_city, top_ten_quantity

    @staticmethod
    def print_statistic(stats1, stats2, stats3, stats4, stats5, stats6):
        """
        Выводит значение статистики на консоль

        :param
            stats1: Динамика уровня зарплат по годам
        :param
            stats2: Динамика количества вакансий по годам
        :param
            stats3: Динамика уровня зарплат по годам для выбранной профессии
        :param
            stats4: Динамика количества вакансий по годам для выбранной профессии
        :param
            stats5: Уровень зарплат по городам (в порядке убывания)
        :param
            stats6: Доля вакансий по городам (в порядке убывания)
        :return:
            void
        """
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))


class InputConnect:
    """
    Класс, который принимает входные данные.
        Atributes:
            file_name(str): название файла
            vacancy_name(str): интирисующая профессия

    """
    def __init__(self):
        """
            Инициализирует объект InputConnect, вызывает методы для подсчета статистики,
            вызывает методы для вывода статистики на консоль,создает объект класса report
        """
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')
        flag = input("Введите метод представления данных: ")
        dataset = DataSet(self.file_name, self.vacancy_name)
        salary_year = dataset.calculate_statistics()[0]
        vacancies_year = dataset.calculate_statistics()[1]
        vacancies_salary_name = dataset.calculate_statistics()[2]
        vacancies_quantity = dataset.calculate_statistics()[3]
        salary_city_decrease = dataset.calculate_statistics()[4]
        salary_city_increase = dataset.calculate_statistics()[5]

        dataset.print_statistic(salary_year,
                                vacancies_year,
                                vacancies_salary_name,
                                vacancies_quantity,
                                salary_city_decrease,
                                salary_city_increase)

        report = Report(self.vacancy_name,
                        salary_year,
                        vacancies_year,
                        vacancies_salary_name,
                        vacancies_quantity,
                        salary_city_decrease,
                        salary_city_increase)
        if flag == "Вакансии":
            report.generate_excel()
            report.save('report.xlsx')
        elif flag == "Статистика":
            report.generate_image()



class Report:
    """
        Класс report создает таблицу excel с данными по интересующей профессии и графики со статистикой
            Attributes:
                wb(Workbook): лист excel
                vacancy_name(Any): название профессии
                stats1(Any): Динамика уровня зарплат по годам
                stats2(Any): Динамика количества вакансий по годам
                stats3(Any): Динамика уровня зарплат по годам для выбранной профессии
                stats4(Any): Динамика количества вакансий по годам для выбранной профессии
                stats5(Any): Уровень зарплат по городам (в порядке убывания)
                stats6(Any): Доля вакансий по городам (в порядке убывания)

    """
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        """
            Инициализирует объект Report
        :param
            vacancy_name(Any): интересующая профессия
        :param
            stats1(Any): Динамика уровня зарплат по годам
        :param
            stats2(Any): Динамика количества вакансий по годам
        :param
            stats3(Any): Динамика уровня зарплат по годам для выбранной профессии
        :param
            stats4(Any): Динамика количества вакансий по годам для выбранной профессии
        :param
            stats5(Any): Уровень зарплат по городам (в порядке убывания)
        :param
            stats6(Any): Доля вакансий по городам (в порядке убывания)
        """
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def generate_excel(self):
        """
            генерирует excel файл с данными по профессии
        :return:
            void
        """
        ws1 = self.wb.active
        ws1.title = 'Статистика по годам'
        ws1.append(['Год',
                    'Средняя зарплата',
                    'Средняя зарплата - ' + self.vacancy_name,
                    'Количество вакансий',
                    'Количество вакансий - ' + self.vacancy_name])

        for year in self.stats1.keys():
            ws1.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        data = [['Год ', 'Средняя зарплата ',
                 ' Средняя зарплата - ' + self.vacancy_name,
                 ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)

        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            ws1[col + '1'].font = font_bold
            ws2[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def generate_image(self):
        """
            вызвает методы по созданию графиков, сохраняет полученное изображение в формате png
        :return:
            void
        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        self.first_diagram(ax1)
        self.second_diagram(ax2)
        self.horizontal_diagram(ax3)
        self.round_diagram(ax4)

        plt.tight_layout()
        plt.savefig('graph.png')

    def round_diagram(self, ax4):
        """
        создает диаграмму - пирог доли вакансий по городам
        :param ax4(axes): диаграмма - пирог
        :return:
            void
        """
        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats6.values()])
        ax4.pie(list(self.stats6.values()) + [other], labels=list(self.stats6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

    def horizontal_diagram(self, ax3):
        """
        создает горизонтальную диаграмму уровень зарплат по городам
        :param ax3(axes): горизонтальная диаграмма
        :return:
            void
        """
        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats5.keys()))]),
                 list(reversed(list(self.stats5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

    def second_diagram(self, ax2):
        """
        создает  диаграмму, показывающую количество вакансий по годам
        :param ax2(axes): диаграмма
        :return:
            void
        """
        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.stats2.keys())) - 0.4, self.stats2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.stats2.keys())), self.stats4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.stats2.keys())) - 0.2, list(self.stats2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

    def first_diagram(self, ax1):
        """
        создает горизонтальную диаграмму, показывающую уровень зарплат по годам
        :param ax4(axes):диаграмма
        :return:
            void
        """
        bar1 = ax1.bar(np.array(list(self.stats1.keys())) - 0.4, self.stats1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.stats1.keys())), self.stats3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.stats1.keys())) - 0.2, list(self.stats1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

    def save(self, filename):
        self.wb.save(filename=filename)


class Salary:
    """
    Класс для представления зарплаты
        Attributes:
            vacancy(dict): интересующая вакансия
    """
    currency_in_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        """
        Инициализирует объект Salary, выполняет конвертацию валюты в рубли, считает среднее значение зарплаты в рублях
        :param vacancy: интересующая вакансия
        """
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        salary_rub = self.currency_in_rub[self.salary_currency]
        salary_average_value = (self.salary_from + self.salary_to) / 2
        self.salary_average = salary_rub * salary_average_value
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


if __name__ == '__main__':
    InputConnect()